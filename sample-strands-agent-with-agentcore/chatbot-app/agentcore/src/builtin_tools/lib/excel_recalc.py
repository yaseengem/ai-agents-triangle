"""
Excel Formula Recalculation using LibreOffice.

Recalculates all formulas in an Excel file and scans for errors.
Runs on the AgentCore runtime container where LibreOffice is available.
"""

import logging
import os
import platform
import subprocess
import tempfile
from pathlib import Path

logger = logging.getLogger(__name__)

MACRO_DIR_MACOS = "~/Library/Application Support/LibreOffice/4/user/basic/Standard"
MACRO_DIR_LINUX = "~/.config/libreoffice/4/user/basic/Standard"
MACRO_FILENAME = "Module1.xba"

RECALCULATE_MACRO = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>"""

EXCEL_ERRORS = ["#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"]


def _setup_macro() -> bool:
    """Set up LibreOffice RecalculateAndSave macro (one-time)."""
    macro_dir = os.path.expanduser(
        MACRO_DIR_MACOS if platform.system() == "Darwin" else MACRO_DIR_LINUX
    )
    macro_file = os.path.join(macro_dir, MACRO_FILENAME)

    if os.path.exists(macro_file):
        try:
            if "RecalculateAndSave" in Path(macro_file).read_text():
                return True
        except Exception:
            pass

    if not os.path.exists(macro_dir):
        try:
            subprocess.run(
                ["soffice", "--headless", "--terminate_after_init"],
                capture_output=True,
                timeout=15,
            )
        except Exception as e:
            logger.warning(f"LibreOffice profile init failed: {e}")
        os.makedirs(macro_dir, exist_ok=True)

    try:
        Path(macro_file).write_text(RECALCULATE_MACRO)
        logger.info(f"LibreOffice recalc macro installed: {macro_file}")
        return True
    except Exception as e:
        logger.error(f"Failed to install LibreOffice macro: {e}")
        return False


def _scan_errors(filename: str) -> dict:
    """Scan recalculated file for Excel formula errors and count formulas."""
    from openpyxl import load_workbook

    error_details = {err: [] for err in EXCEL_ERRORS}
    total_errors = 0
    formula_count = 0

    try:
        wb = load_workbook(filename, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str):
                        for err in EXCEL_ERRORS:
                            if err in cell.value:
                                error_details[err].append(f"{sheet_name}!{cell.coordinate}")
                                total_errors += 1
                                break
        wb.close()

        wb_formulas = load_workbook(filename, data_only=False)
        for sheet_name in wb_formulas.sheetnames:
            ws = wb_formulas[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_count += 1
        wb_formulas.close()

    except Exception as e:
        logger.error(f"Error scanning spreadsheet: {e}")
        return {"status": "scan_error", "error": str(e)}

    result = {
        "status": "success" if total_errors == 0 else "errors_found",
        "total_errors": total_errors,
        "total_formulas": formula_count,
    }

    if total_errors > 0:
        result["error_summary"] = {}
        for err_type, locations in error_details.items():
            if locations:
                result["error_summary"][err_type] = {
                    "count": len(locations),
                    "locations": locations[:10],
                }

    return result


def recalc_spreadsheet(
    file_bytes: bytes,
    filename: str = "temp.xlsx",
    timeout: int = 30
) -> tuple[bytes, dict]:
    """Recalculate formulas in an Excel file using LibreOffice.

    Args:
        file_bytes: Excel file contents as bytes
        filename: Original filename (for temp file naming)
        timeout: LibreOffice timeout in seconds

    Returns:
        (recalculated_bytes, report) where report has:
        - status: "success" | "errors_found" | "skipped"
        - total_formulas: number of formulas
        - total_errors: number of formula errors
        - error_summary: error types and locations (if errors found)
    """
    if not _setup_macro():
        logger.warning("LibreOffice macro setup failed, skipping recalc")
        return file_bytes, {"status": "skipped", "reason": "macro_setup_failed"}

    with tempfile.TemporaryDirectory() as temp_dir:
        temp_path = os.path.join(temp_dir, filename)

        with open(temp_path, 'wb') as f:
            f.write(file_bytes)

        abs_path = str(Path(temp_path).absolute())
        cmd = [
            "soffice",
            "--headless",
            "--norestore",
            "vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application",
            abs_path,
        ]

        if platform.system() == "Linux":
            cmd = ["timeout", str(timeout)] + cmd

        try:
            result = subprocess.run(
                cmd, capture_output=True, text=True, timeout=timeout + 10
            )
            if result.returncode != 0 and result.returncode != 124:
                error_msg = result.stderr or "Unknown recalculation error"
                logger.warning(f"LibreOffice recalc failed (rc={result.returncode}): {error_msg[:200]}")
                return file_bytes, {"status": "skipped", "reason": f"recalc_failed"}
        except subprocess.TimeoutExpired:
            logger.warning(f"LibreOffice recalc timed out after {timeout}s")
            return file_bytes, {"status": "skipped", "reason": "timeout"}
        except FileNotFoundError:
            logger.warning("soffice not found, skipping recalc")
            return file_bytes, {"status": "skipped", "reason": "soffice_not_found"}

        try:
            with open(temp_path, 'rb') as f:
                recalced_bytes = f.read()
        except Exception as e:
            logger.error(f"Failed to read recalculated file: {e}")
            return file_bytes, {"status": "skipped", "reason": f"read_failed"}

        report = _scan_errors(temp_path)
        return recalced_bytes, report
