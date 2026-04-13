"""
Excel Spreadsheet Tools - 5 essential tools for Excel spreadsheet management.

Tools:
1. create_excel_spreadsheet - Create new Excel spreadsheet from Python code
2. modify_excel_spreadsheet - Modify existing Excel spreadsheet with openpyxl code
3. list_my_excel_spreadsheets - List all Excel spreadsheets in workspace
4. read_excel_spreadsheet - Retrieve spreadsheet for download
5. preview_excel_sheets - Get sheet screenshot for visual inspection

Note: Uploaded .xlsx files are automatically stored to workspace by agent.py
Pattern follows word_document_tool for Code Interpreter usage.
"""

import os
import re
import logging
from typing import Dict, Any, Optional
from strands import tool, ToolContext
from skill import register_skill
from workspace import ExcelManager
from builtin_tools.lib.excel_recalc import recalc_spreadsheet
from builtin_tools.lib.tool_response import build_success_response, build_image_response

logger = logging.getLogger(__name__)


def _format_recalc_report(report: dict) -> str:
    """Format recalc report as a message string for the response."""
    status = report.get("status", "skipped")
    if status == "skipped":
        return ""

    formulas = report.get("total_formulas", 0)
    if formulas == 0:
        return ""

    parts = [f"**Formulas**: {formulas} recalculated"]

    if status == "errors_found":
        errors = report.get("total_errors", 0)
        parts.append(f"**Formula Errors**: {errors} found")
        for err_type, info in report.get("error_summary", {}).items():
            locations = ", ".join(info["locations"][:5])
            parts.append(f"  - {err_type} ({info['count']}): {locations}")

    return "\n".join(parts)


def _validate_spreadsheet_name(name: str) -> tuple[bool, Optional[str]]:
    """Validate spreadsheet name meets requirements (without extension).

    Rules:
    - Only letters (a-z, A-Z), numbers (0-9), hyphens (-), and underscores (_)
    - No spaces or special characters
    - No consecutive hyphens
    - No leading/trailing hyphens

    Args:
        name: Spreadsheet name without extension (e.g., "sales-report")

    Returns:
        (is_valid, error_message)
        - (True, None) if valid
        - (False, error_message) if invalid
    """
    # Check for empty name
    if not name:
        return False, "Spreadsheet name cannot be empty"

    # Check for valid characters: letters, numbers, hyphens, underscores
    if not re.match(r'^[a-zA-Z0-9_\-]+$', name):
        invalid_chars = re.findall(r'[^a-zA-Z0-9_\-]', name)
        return False, f"Invalid characters in name: {set(invalid_chars)}. Use only letters, numbers, hyphens, and underscores."

    # Check for consecutive hyphens
    if '--' in name:
        return False, "Name cannot contain consecutive hyphens (--)"

    # Check for leading/trailing hyphens
    if name.startswith('-') or name.endswith('-'):
        return False, "Name cannot start or end with a hyphen"

    return True, None


def _sanitize_spreadsheet_name_for_bedrock(filename: str) -> str:
    """Sanitize existing filename for Bedrock API (removes extension).

    Use this ONLY for existing files being read from S3.
    For new files, use _validate_spreadsheet_name() instead.

    Args:
        filename: Original filename with extension (e.g., "test_spreadsheet_v2.xlsx")

    Returns:
        Sanitized name without extension (e.g., "test-spreadsheet-v2")
    """
    # Remove extension
    if '.' in filename:
        name, ext = filename.rsplit('.', 1)
    else:
        name = filename

    # Replace underscores and spaces with hyphens
    name = name.replace('_', '-').replace(' ', '-')

    # Keep only allowed characters: alphanumeric, hyphens, parentheses, square brackets
    # This matches agent.py's _sanitize_filename behavior
    name = re.sub(r'[^a-zA-Z0-9\-\(\)\[\]]', '', name)

    # Replace multiple consecutive hyphens with single hyphen
    name = re.sub(r'\-+', '-', name)

    # Trim hyphens from start/end
    name = name.strip('-')

    # If name becomes empty, use default
    if not name:
        name = 'spreadsheet'

    if name != filename.replace('.xlsx', ''):
        logger.info(f"Sanitized spreadsheet name for Bedrock: '{filename}' → '{name}'")

    return name


def _get_user_session_ids(tool_context: ToolContext) -> tuple[str, str]:
    """Extract user_id and session_id from ToolContext

    Returns:
        (user_id, session_id) tuple
    """
    # Extract from invocation_state (set by agent)
    invocation_state = tool_context.invocation_state
    user_id = invocation_state.get('user_id', 'default_user')
    session_id = invocation_state.get('session_id', 'default_session')

    logger.info(f"Extracted IDs: user_id={user_id}, session_id={session_id}")
    return user_id, session_id


def _save_excel_artifact(
    tool_context: ToolContext,
    filename: str,
    s3_url: str,
    size_kb: str,
    tool_name: str,
    user_id: str,
    session_id: str
) -> None:
    """Save Excel spreadsheet as artifact to agent.state for Canvas display.

    Args:
        tool_context: Strands ToolContext
        filename: Spreadsheet filename (e.g., "report.xlsx")
        s3_url: Full S3 URL (e.g., "s3://bucket/path/report.xlsx")
        size_kb: File size string (e.g., "45.2 KB")
        tool_name: Tool that created this ("create_excel_spreadsheet" or "modify_excel_spreadsheet")
        user_id: User ID
        session_id: Session ID
    """
    from datetime import datetime, timezone

    try:
        # Generate artifact ID using filename (without extension) for easy lookup
        sheet_name = filename.replace('.xlsx', '')
        artifact_id = f"excel-{sheet_name}"

        # Get current artifacts from agent.state
        artifacts = tool_context.agent.state.get("artifacts") or {}

        # Create/update artifact
        artifacts[artifact_id] = {
            "id": artifact_id,
            "type": "excel_spreadsheet",
            "title": filename,
            "content": s3_url,  # Full S3 URL for OfficeViewer
            "tool_name": tool_name,
            "metadata": {
                "filename": filename,
                "s3_url": s3_url,
                "size_kb": size_kb,
                "user_id": user_id,
                "session_id": session_id
            },
            "created_at": artifacts.get(artifact_id, {}).get("created_at", datetime.now(timezone.utc).isoformat()),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }

        # Save to agent.state
        tool_context.agent.state.set("artifacts", artifacts)

        # Sync agent state to persistence
        session_manager = tool_context.invocation_state.get("session_manager")
        if not session_manager and hasattr(tool_context.agent, 'session_manager'):
            session_manager = tool_context.agent.session_manager

        if session_manager:
            session_manager.sync_agent(tool_context.agent)
            logger.info(f"Saved Excel artifact: {artifact_id}")
        else:
            logger.warning(f"No session_manager found, Excel artifact not persisted: {artifact_id}")

    except Exception as e:
        logger.error(f"Failed to save Excel artifact: {e}")


@tool(context=True)
def create_excel_spreadsheet(
    python_code: str,
    spreadsheet_name: str,
    tool_context: ToolContext
) -> Dict[str, Any]:
    """Create a new Excel spreadsheet using openpyxl code.

    This tool executes openpyxl code to create a spreadsheet from scratch.
    Perfect for generating structured data with sheets, tables, charts, and formatting.

    Available libraries: openpyxl, pandas, matplotlib, numpy

    Args:
        python_code: Python code using openpyxl to build the spreadsheet.
                    The workbook is initialized as: wb = Workbook()
                    The active sheet is: ws = wb.active
                    After your code, it's automatically saved.

                    DO NOT include Workbook() initialization or wb.save() calls.

                    Uploaded images are automatically available in Code Interpreter.
                    Use os.listdir() to discover available image files.

                    Common Patterns:

                    Basic Data Entry:
                    ```python
# Set sheet title
ws.title = 'Sales Data'

# Add headers with formatting
ws['A1'] = 'Product'
ws['B1'] = 'Quantity'
ws['C1'] = 'Price'
from openpyxl.styles import Font, PatternFill
ws['A1'].font = Font(bold=True, size=12)
ws['A1'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

# Add data rows
data = [
    ['Widget A', 100, 25.50],
    ['Widget B', 150, 30.00],
    ['Widget C', 200, 20.00]
]
for row in data:
    ws.append(row)
                    ```

                    With Table:
                    ```python
from openpyxl.worksheet.table import Table, TableStyleInfo

# Add data first
ws['A1'] = 'Product'
ws['B1'] = 'Sales'
ws.append(['Product A', 1000])
ws.append(['Product B', 1500])

# Create table
tab = Table(displayName='SalesTable', ref='A1:B3')
style = TableStyleInfo(name='TableStyleMedium9', showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
tab.tableStyleInfo = style
ws.add_table(tab)
                    ```

                    With Chart:
                    ```python
from openpyxl.chart import BarChart, Reference

# Add data
ws['A1'] = 'Month'
ws['B1'] = 'Sales'
for i, (month, sales) in enumerate([('Jan', 100), ('Feb', 120), ('Mar', 150)], 2):
    ws[f'A{i}'] = month
    ws[f'B{i}'] = sales

# Create chart
chart = BarChart()
chart.title = 'Monthly Sales'
chart.x_axis.title = 'Month'
chart.y_axis.title = 'Sales'

data = Reference(ws, min_col=2, min_row=1, max_row=4)
categories = Reference(ws, min_col=1, min_row=2, max_row=4)
chart.add_data(data, titles_from_data=True)
chart.set_categories(categories)

ws.add_chart(chart, 'D2')
                    ```

                    With Uploaded Image:
                    ```python
from openpyxl.drawing.image import Image
import os

# Discover available images
available_images = [f for f in os.listdir() if f.endswith(('.png', '.jpg', '.jpeg', '.gif', '.webp'))]

if available_images:
    # Use the first image (or select specific one by filename matching)
    image_file = available_images[0]
    img = Image(image_file)
    ws.add_image(img, 'E1')
else:
    # No images found - add note in cell
    ws['E1'] = '[No images available in workspace]'
                    ```

                    With Multiple Sheets:
                    ```python
# Create additional sheets
ws2 = wb.create_sheet('Summary')
ws3 = wb.create_sheet('Details')

# Add data to each sheet
ws.title = 'Overview'
ws['A1'] = 'Main Data'

ws2['A1'] = 'Summary'
ws3['A1'] = 'Detailed Analysis'
                    ```

        spreadsheet_name: Spreadsheet name WITHOUT extension (.xlsx is added automatically)
                         Use ONLY letters, numbers, hyphens (no underscores or spaces)
                         Examples: "sales-report", "Q4-data", "inventory-2024"

    Returns:
        Success message with file details and workspace list

    Note:
        - Spreadsheet is saved to workspace for future editing
        - Uploaded images are automatically available in Code Interpreter
        - Keep code focused on structure; use modify_excel_spreadsheet for refinements
    """
    try:
        logger.info("=== create_excel_spreadsheet called ===")
        logger.info(f"Spreadsheet name: {spreadsheet_name}")

        # Validate spreadsheet name (without extension)
        is_valid, error_msg = _validate_spreadsheet_name(spreadsheet_name)
        if not is_valid:
            return {
                "content": [{
                    "text": f"**Invalid spreadsheet name**: {spreadsheet_name}\n\n{error_msg}\n\n**Examples of valid names:**\n- sales-report\n- Q4-data\n- inventory-2024"
                }],
                "status": "error"
            }

        # Add .xlsx extension
        spreadsheet_filename = f"{spreadsheet_name}.xlsx"
        logger.info(f"Full filename: {spreadsheet_filename}")

        # Get user and session IDs
        user_id, session_id = _get_user_session_ids(tool_context)

        # Initialize document manager
        doc_manager = ExcelManager(user_id, session_id)

        # Get shared CI session
        from builtin_tools.code_interpreter_tool import get_ci_session
        code_interpreter = get_ci_session(tool_context)
        if code_interpreter is None:
            return {
                "content": [{
                    "text": "**Code Interpreter not configured**\n\nCODE_INTERPRETER_ID not found in environment or Parameter Store."
                }],
                "status": "error"
            }

        try:
            # Load all workspace images from S3 to Code Interpreter
            loaded_images = doc_manager.load_workspace_images_to_ci(code_interpreter)
            if loaded_images:
                logger.info(f"Loaded {len(loaded_images)} image(s) from workspace: {loaded_images}")

            # Get Code Interpreter path for file
            ci_path = doc_manager.get_ci_path(spreadsheet_filename)

            # Build spreadsheet creation code
            creation_code = f"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.worksheet.table import Table, TableStyleInfo

# Create new workbook
wb = Workbook()
ws = wb.active

# Execute user's creation code
{python_code}

# Save workbook
wb.save('{ci_path}')
print(f"Spreadsheet created: {ci_path}")
"""

            # Execute creation
            response = code_interpreter.invoke("executeCode", {
                "code": creation_code,
                "language": "python",
                "clearContext": False
            })

            # Check for errors and capture stdout
            stdout_output = ""
            for event in response.get("stream", []):
                result = event.get("result", {})
                if result.get("isError", False):
                    error_msg = result.get("structuredContent", {}).get("stderr", "Unknown error")
                    logger.error(f"Creation failed: {error_msg[:500]}")
                    return {
                        "content": [{
                            "text": f"**Failed to create spreadsheet**\n\n```\n{error_msg[:1000]}\n```\n\nTip:Check your openpyxl code for syntax errors or incorrect API usage."
                        }],
                        "status": "error"
                    }
                # Capture stdout
                stdout = result.get("structuredContent", {}).get("stdout", "")
                if stdout:
                    stdout_output += stdout

            logger.info("Spreadsheet creation completed")

            # Download from Code Interpreter
            file_bytes = doc_manager.download_from_code_interpreter(code_interpreter, spreadsheet_filename)

            # Recalculate formulas using LibreOffice
            file_bytes, recalc_report = recalc_spreadsheet(file_bytes, spreadsheet_filename)
            recalc_msg = _format_recalc_report(recalc_report)

            # Save to S3 for persistence
            s3_info = doc_manager.save_to_s3(
                spreadsheet_filename,
                file_bytes,
                metadata={'source': 'python_code_creation'}
            )

            # Save as artifact for Canvas display
            _save_excel_artifact(
                tool_context=tool_context,
                filename=spreadsheet_filename,
                s3_url=s3_info['s3_url'],
                size_kb=s3_info['size_kb'],
                tool_name='create_excel_spreadsheet',
                user_id=user_id,
                session_id=session_id
            )

            # Get current workspace list
            workspace_docs = doc_manager.list_s3_documents()
            other_files_count = len([d for d in workspace_docs if d['filename'] != spreadsheet_filename])

            message = f"""**Spreadsheet created successfully**

**File**: {spreadsheet_filename} ({s3_info['size_kb']})
**Other files in workspace**: {other_files_count} spreadsheet{'s' if other_files_count != 1 else ''}"""

            if recalc_msg:
                message += f"\n{recalc_msg}"

            # Include stdout output if any
            if stdout_output.strip():
                message += f"\n\n**Output:**\n```\n{stdout_output.strip()}\n```"

            # Return success message
            return build_success_response(message, {
                "filename": spreadsheet_filename,
                "tool_type": "excel_spreadsheet",
                "user_id": user_id,
                "session_id": session_id
            })

        except Exception as e:
            logger.error(f"CI execution error in create_excel_spreadsheet: {e}")
            raise

    except Exception as e:
        logger.error(f"create_excel_spreadsheet failed: {e}")
        return {
            "content": [{
                "text": f"**Failed to create spreadsheet**\n\n{str(e)}"
            }],
            "status": "error"
        }


@tool(context=True)
def modify_excel_spreadsheet(
    source_name: str,
    output_name: str,
    python_code: str,
    tool_context: ToolContext
) -> Dict[str, Any]:
    """Modify existing Excel spreadsheet using openpyxl code and save with a new name.

    This tool loads a spreadsheet from workspace, executes openpyxl code to modify it,
    and saves it with a new filename to preserve the original.

    Available libraries: openpyxl, pandas, matplotlib, numpy

    IMPORTANT Safety Rules:
    - Always use different output_name than source_name (e.g., "report" → "report-v2")
    - Always check sheet exists before accessing
    - Use try-except for operations that might fail

    Args:
        source_name: Spreadsheet name to load (WITHOUT extension, must exist in workspace)
                    Example: "sales-report", "Q4-data"
        output_name: New spreadsheet name (WITHOUT extension, must be different from source)
                    Use ONLY letters, numbers, hyphens (no underscores or spaces)
                    Example: "sales-report-v2", "Q4-data-final"
        python_code: Python code using openpyxl library to modify spreadsheet.
                    The workbook is loaded as: wb = load_workbook('<filename>')
                    After modifications, it's automatically saved.

                    DO NOT include load_workbook() or wb.save() calls.

                    IMPORTANT: Uploaded images are automatically available in Code Interpreter.

                    Common Patterns:

                    Add New Sheet with Data:
                    ```python
# Create new sheet
ws_new = wb.create_sheet('Q1 Summary')

# Add data
ws_new['A1'] = 'Summary Data'
ws_new['A1'].font = Font(bold=True, size=14)
ws_new.append(['Item', 'Value'])
ws_new.append(['Total', 10000])
                    ```

                    Modify Existing Data:
                    ```python
# Access existing sheet
ws = wb['Sales Data']

# Update specific cells
ws['B2'] = 150  # Update value
ws['C2'].font = Font(color='FF0000')  # Change color
                    ```

                    Add Chart to Existing Sheet:
                    ```python
from openpyxl.chart import BarChart, Reference

# Access sheet
ws = wb.active

# Create chart from existing data
chart = BarChart()
chart.title = 'Sales Analysis'
data = Reference(ws, min_col=2, min_row=1, max_row=10)
categories = Reference(ws, min_col=1, min_row=2, max_row=10)
chart.add_data(data, titles_from_data=True)
chart.set_categories(categories)

ws.add_chart(chart, 'E2')
                    ```

                    Insert Uploaded Image:
                    ```python
from openpyxl.drawing.image import Image

# Access sheet
ws = wb['Dashboard']

# Images from workspace are automatically available
import os
images = [f for f in os.listdir() if f.endswith(('.png', '.jpg', '.jpeg'))]
if images:
    img = Image(images[0])
    ws.add_image(img, 'F5')
                    ```

    Returns:
        Success message with file details and workspace list

    Note:
        - Uploaded images are automatically available in Code Interpreter
        - Document automatically synced to S3
    """
    try:
        logger.info("=== modify_excel_spreadsheet called ===")
        logger.info(f"Source: {source_name}, Output: {output_name}")

        # Validate output name format
        is_valid, error_msg = _validate_spreadsheet_name(output_name)
        if not is_valid:
            return {
                "content": [{
                    "text": f"**Invalid output name**: {output_name}\n\n{error_msg}\n\n**Examples of valid names:**\n- sales-report-v2\n- Q4-data-final\n- report-revised"
                }],
                "status": "error"
            }

        # Ensure source and output are different
        if source_name == output_name:
            return {
                "content": [{
                    "text": f"**Invalid name**\n\nOutput name must be different from source name to preserve the original.\n\nSource: {source_name}\nOutput: {output_name}\n\nTip:Try: \"{source_name}-v2\""
                }],
                "status": "error"
            }

        # Add .xlsx extensions
        source_filename = f"{source_name}.xlsx"
        output_filename = f"{output_name}.xlsx"
        logger.info(f"Full filenames: {source_filename} → {output_filename}")

        # Get user and session IDs
        user_id, session_id = _get_user_session_ids(tool_context)

        # Initialize document manager
        doc_manager = ExcelManager(user_id, session_id)

        # Get shared CI session
        from builtin_tools.code_interpreter_tool import get_ci_session
        code_interpreter = get_ci_session(tool_context)
        if code_interpreter is None:
            return {
                "content": [{
                    "text": "**Code Interpreter not configured**\n\nCODE_INTERPRETER_ID not found in environment or Parameter Store."
                }],
                "status": "error"
            }

        try:
            # Load all workspace images from S3 to Code Interpreter
            loaded_images = doc_manager.load_workspace_images_to_ci(code_interpreter)
            if loaded_images:
                logger.info(f"Loaded {len(loaded_images)} image(s) from workspace: {loaded_images}")

            # Ensure source file is in Code Interpreter (load from S3 if needed)
            source_ci_path = doc_manager.ensure_file_in_ci(code_interpreter, source_filename)

            # Generate output path
            output_ci_path = doc_manager.get_ci_path(output_filename)

            # Build modification code
            modification_code = f"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.worksheet.table import Table, TableStyleInfo

# Load source spreadsheet
wb = load_workbook('{source_ci_path}')

# Execute user's modification code
{python_code}

# Save to output file
wb.save('{output_ci_path}')
print(f"Spreadsheet modified and saved: {output_ci_path}")
"""

            # Execute modification
            response = code_interpreter.invoke("executeCode", {
                "code": modification_code,
                "language": "python",
                "clearContext": False
            })

            # Check for errors and capture stdout
            stdout_output = ""
            for event in response.get("stream", []):
                result = event.get("result", {})
                if result.get("isError", False):
                    error_msg = result.get("structuredContent", {}).get("stderr", "Unknown error")
                    logger.error(f"Modification failed: {error_msg[:500]}")
                    return {
                        "content": [{
                            "text": f"**Modification failed**\n\n```\n{error_msg[:1000]}\n```\n\nTip:Check your openpyxl code for syntax errors or incorrect API usage."
                        }],
                        "status": "error"
                    }
                # Capture stdout
                stdout = result.get("structuredContent", {}).get("stdout", "")
                if stdout:
                    stdout_output += stdout

            logger.info("Spreadsheet modification completed")

            # Download modified spreadsheet from Code Interpreter
            file_bytes = doc_manager.download_from_code_interpreter(code_interpreter, output_filename)

            # Recalculate formulas using LibreOffice
            file_bytes, recalc_report = recalc_spreadsheet(file_bytes, output_filename)
            recalc_msg = _format_recalc_report(recalc_report)

            # Save to S3 with output filename
            s3_info = doc_manager.save_to_s3(
                output_filename,
                file_bytes,
                metadata={
                    'source': 'modification',
                    'source_filename': source_filename,
                    'modified_at': 'timestamp'
                }
            )

            # Save as artifact for Canvas display
            _save_excel_artifact(
                tool_context=tool_context,
                filename=output_filename,
                s3_url=s3_info['s3_url'],
                size_kb=s3_info['size_kb'],
                tool_name='modify_excel_spreadsheet',
                user_id=user_id,
                session_id=session_id
            )

            # Get current workspace list
            workspace_docs = doc_manager.list_s3_documents()
            other_files_count = len([d for d in workspace_docs if d['filename'] != output_filename])

            # Build success message
            message = f"""**Spreadsheet modified successfully**

**Source**: {source_filename}
**Saved as**: {output_filename} ({s3_info['size_kb']})
**Other files in workspace**: {other_files_count} spreadsheet{'s' if other_files_count != 1 else ''}"""

            if recalc_msg:
                message += f"\n{recalc_msg}"

            # Include stdout output if any
            if stdout_output.strip():
                message += f"\n\n**Output:**\n```\n{stdout_output.strip()}\n```"

            # Return success message with metadata for download button
            return build_success_response(message, {
                "filename": output_filename,
                "tool_type": "excel_spreadsheet",
                "user_id": user_id,
                "session_id": session_id
            })

        except Exception as e:
            logger.error(f"CI execution error in modify_excel_spreadsheet: {e}")
            raise

    except FileNotFoundError as e:
        logger.error(f"Spreadsheet not found: {e}")
        return {
            "content": [{
                "text": f"**Spreadsheet not found**: {source_filename}"
            }],
            "status": "error"
        }
    except Exception as e:
        logger.error(f"modify_excel_spreadsheet failed: {e}")
        return {
            "content": [{
                "text": f"**Failed to modify spreadsheet**\n\n{str(e)}"
            }],
            "status": "error"
        }


@tool(context=True)
def list_my_excel_spreadsheets(
    tool_context: ToolContext
) -> Dict[str, Any]:
    """List all Excel spreadsheets in workspace.

    Shows all .xlsx files in workspace with size and metadata.

    Use this tool when:
    - User asks "what Excel files do I have?"
    - User says "show my spreadsheets", "list files"
    - Before modifying: verify spreadsheet exists
    - User wants to see workspace contents

    No arguments needed.

    Returns:
        - Formatted list of all Excel spreadsheets
        - Each entry shows: filename, size, last modified date
        - Total file count
        - Metadata for frontend download buttons

    Example Usage:
        Scenario 1 - Check available files:
            User: "What Excel spreadsheets do I have?"
            AI: list_my_excel_spreadsheets()
            → Shows: sales.xlsx, inventory.xlsx, report.xlsx

        Scenario 2 - Before modifying:
            User: "Edit my sales data"
            AI: [Unclear which file]
            AI: list_my_excel_spreadsheets()
            AI: "I found these spreadsheets: ... Which one should I modify?"

    Example Output:
        Workspace (3 spreadsheets):
          - sales-report.xlsx (52.3 KB) - Modified: 2025-01-15
          - inventory.xlsx (41.8 KB) - Modified: 2025-01-14
          - Q4-analysis.xlsx (89.2 KB) - Modified: 2025-01-13

    Note:
        - Shows files from workspace
        - Empty workspace shows helpful message
        - Frontend renders download buttons automatically
    """
    try:
        logger.info("=== list_my_excel_spreadsheets called ===")

        # Get user and session IDs
        user_id, session_id = _get_user_session_ids(tool_context)

        # Initialize document manager
        doc_manager = ExcelManager(user_id, session_id)

        # List documents from S3
        documents = doc_manager.list_s3_documents()

        # Format list
        workspace_summary = doc_manager.format_file_list(documents)

        if documents:
            message = workspace_summary
        else:
            message = workspace_summary

        # Prepare metadata for frontend (download buttons)
        metadata = {
            "documents": [
                {
                    "filename": doc['filename'],
                    "s3_key": doc['s3_key'],
                    "size_kb": doc['size_kb'],
                    "last_modified": doc['last_modified']
                } for doc in documents
            ]
        }

        return build_success_response(message, metadata)

    except Exception as e:
        logger.error(f"list_my_excel_spreadsheets failed: {e}")
        return {
            "content": [{
                "text": f"**Failed to list spreadsheets**\n\n{str(e)}"
            }],
            "status": "error"
        }


@tool(context=True)
def read_excel_spreadsheet(
    spreadsheet_name: str,
    tool_context: ToolContext
) -> Dict[str, Any]:
    """Read and retrieve a specific Excel spreadsheet.

    This tool loads a spreadsheet from workspace and extracts its data content using Code Interpreter.
    The extracted data (sheets, rows, columns) is returned for analysis and answering questions.

    Use this tool when:
    - User asks about spreadsheet contents: "What's in sales.xlsx?", "Summarize this data"
    - User wants to analyze the spreadsheet: "How many sheets?", "What's the total?"
    - User explicitly requests download: "Send me [filename]", "I need [spreadsheet]"
    - You need to verify spreadsheet contents before modification

    IMPORTANT:
    - For creating new spreadsheets: use create_excel_spreadsheet
    - For modifying spreadsheets: use modify_excel_spreadsheet

    Args:
        spreadsheet_name: Spreadsheet name WITHOUT extension (.xlsx is added automatically)
                         Must exist in workspace.
                         Example: "sales-report", "inventory", "Q4-data"

    Returns:
        - Extracted data content (sheet names, cell values, table format)
        - Spreadsheet metadata (filename, size, S3 location)
        - Frontend shows download button based on metadata

    Example Usage:
        # Download request
        User: "Send me the sales report"
        AI: read_excel_spreadsheet("sales-report")

        # After creation
        User: "Create report and send it"
        AI: create_excel_spreadsheet(...)
        AI: read_excel_spreadsheet("sales-report")

    Note:
        - File must exist in workspace
        - Frontend handles download automatically
    """
    try:
        logger.info("=== read_excel_spreadsheet called ===")
        logger.info(f"Spreadsheet name: {spreadsheet_name}")

        # Add .xlsx extension
        spreadsheet_filename = f"{spreadsheet_name}.xlsx"
        logger.info(f"Full filename: {spreadsheet_filename}")

        # Get user and session IDs
        user_id, session_id = _get_user_session_ids(tool_context)

        # Initialize document manager
        doc_manager = ExcelManager(user_id, session_id)

        # Load from S3
        file_bytes = doc_manager.load_from_s3(spreadsheet_filename)

        # Get file info
        documents = doc_manager.list_s3_documents()
        doc_info = next((d for d in documents if d['filename'] == spreadsheet_filename), None)

        if not doc_info:
            raise FileNotFoundError(f"Spreadsheet not found: {spreadsheet_filename}")

        # Get shared CI session
        from builtin_tools.code_interpreter_tool import get_ci_session
        code_interpreter = get_ci_session(tool_context)
        if code_interpreter is None:
            return {
                "content": [{
                    "text": "**Code Interpreter not configured**\n\nCODE_INTERPRETER_ID not found in environment or Parameter Store."
                }],
                "status": "error"
            }

        try:
            # Upload spreadsheet to Code Interpreter
            doc_manager.upload_to_code_interpreter(code_interpreter, spreadsheet_filename, file_bytes)

            # Generate extraction code
            extraction_code = f'''
import json
from openpyxl import load_workbook

wb = load_workbook("{spreadsheet_filename}", data_only=True)
result = {{
    "sheets": [],
    "properties": {{
        "sheet_count": len(wb.sheetnames),
        "sheet_names": wb.sheetnames
    }}
}}

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    sheet_data = {{
        "name": sheet_name,
        "rows": sheet.max_row or 0,
        "cols": sheet.max_column or 0,
        "data": []
    }}

    # Limit rows for efficiency
    max_rows = min(sheet.max_row or 0, 100)
    max_cols = min(sheet.max_column or 0, 20)

    for row in sheet.iter_rows(min_row=1, max_row=max_rows, max_col=max_cols, values_only=True):
        row_values = [str(cell) if cell is not None else "" for cell in row]
        if any(v.strip() for v in row_values):  # Skip empty rows
            sheet_data["data"].append(row_values)

    if sheet.max_row and sheet.max_row > 100:
        sheet_data["truncated"] = True
        sheet_data["total_rows"] = sheet.max_row

    result["sheets"].append(sheet_data)

print(json.dumps(result, ensure_ascii=False))
'''

            # Execute extraction
            response = code_interpreter.invoke("executeCode", {
                "code": extraction_code,
                "language": "python",
                "clearContext": False
            })

            # Collect JSON output
            json_output = ""
            for event in response.get("stream", []):
                result = event.get("result", {})
                if result.get("isError", False):
                    error_msg = result.get("structuredContent", {}).get("stderr", "Unknown error")
                    logger.error(f"Extraction failed: {error_msg[:500]}")
                    return {
                        "content": [{
                            "text": f"**Failed to read spreadsheet**\n\n```\n{error_msg[:1000]}\n```"
                        }],
                        "status": "error"
                    }

                stdout = result.get("structuredContent", {}).get("stdout", "")
                if stdout:
                    json_output += stdout

            # Parse JSON result
            import json
            spreadsheet_content = json.loads(json_output)

            # Format output text
            output_parts = []
            props = spreadsheet_content.get("properties", {})
            output_parts.append(f"**Spreadsheet Content**: {spreadsheet_filename} ({doc_info['size_kb']})")
            output_parts.append(f"**Sheets**: {', '.join(props.get('sheet_names', []))}")
            output_parts.append("")

            # Format each sheet
            for sheet in spreadsheet_content.get("sheets", []):
                output_parts.append(f"### Sheet: {sheet['name']} ({sheet['rows']} rows × {sheet['cols']} cols)")
                output_parts.append("")

                data = sheet.get("data", [])
                if data:
                    # Header row
                    if len(data) > 0:
                        output_parts.append(" | ".join(data[0]))
                        output_parts.append(" | ".join(["---"] * len(data[0])))

                    # Data rows
                    for row in data[1:]:
                        output_parts.append(" | ".join(row))

                if sheet.get("truncated"):
                    output_parts.append(f"\n... (showing first 100 of {sheet.get('total_rows', '?')} rows)")

                output_parts.append("")

            # Add summary
            output_parts.append(f"---\n*{props.get('sheet_count', 0)} sheet(s)*")
            output_parts.append(f"*Last Modified: {doc_info['last_modified'].split('T')[0]}*")

            output_text = "\n".join(output_parts)

            # Truncate if too long
            max_chars = 15000
            if len(output_text) > max_chars:
                output_text = output_text[:max_chars] + f"\n\n... (truncated, total {len(output_text)} characters)"

            return build_success_response(output_text, {
                "filename": spreadsheet_filename,
                "s3_key": doc_manager.get_s3_key(spreadsheet_filename),
                "size_kb": doc_info['size_kb'],
                "last_modified": doc_info['last_modified'],
                "tool_type": "excel_spreadsheet",
                "user_id": user_id,
                "session_id": session_id
            })

        except Exception as e:
            logger.error(f"CI execution error in read_excel_spreadsheet: {e}")
            raise

    except FileNotFoundError as e:
        logger.error(f"Spreadsheet not found: {e}")
        return {
            "content": [{
                "text": f"**Spreadsheet not found**: {spreadsheet_filename}"
            }],
            "status": "error"
        }
    except Exception as e:
        logger.error(f"read_excel_spreadsheet failed: {e}")
        return {
            "content": [{
                "text": f"**Failed to read spreadsheet**\n\n{str(e)}"
            }],
            "status": "error"
        }


@tool(context=True)
def preview_excel_sheets(
    spreadsheet_name: str,
    sheet_names: list[str],
    tool_context: ToolContext
) -> Dict[str, Any]:
    """Get sheet screenshots for YOU (the agent) to visually inspect before editing.

    This tool is for YOUR internal use - to see the actual layout, formatting,
    charts, and data of sheets before making modifications. Images are sent to you,
    not displayed to the user.

    Args:
        spreadsheet_name: Spreadsheet name without extension (e.g., "sales-report")
        sheet_names: List of sheet names to preview. Use empty list [] for all sheets.
                    Example: ["Sheet1", "Summary"] or []

    Use BEFORE modifying a spreadsheet to:
    - See exact data layout and formatting
    - Identify charts, images, or conditional formatting
    - Understand column widths and row heights
    - Plan precise edits based on visual layout
    """
    import subprocess
    import tempfile
    import base64
    import io
    from pdf2image import convert_from_path
    from openpyxl import load_workbook

    # Get user and session IDs
    user_id, session_id = _get_user_session_ids(tool_context)

    # Validate and prepare filename
    spreadsheet_filename = f"{spreadsheet_name}.xlsx"
    logger.info(f"preview_excel_sheets: {spreadsheet_filename}, sheets {sheet_names}")

    try:
        # Initialize document manager
        doc_manager = ExcelManager(user_id, session_id)

        # Check if spreadsheet exists
        documents = doc_manager.list_s3_documents()
        doc_info = next((d for d in documents if d['filename'] == spreadsheet_filename), None)

        if not doc_info:
            available = [d['filename'] for d in documents]
            return {
                "content": [{
                    "text": f"**Spreadsheet not found**: {spreadsheet_filename}\n\n"
                           f"Available spreadsheets: {', '.join(available) if available else 'None'}"
                }],
                "status": "error"
            }

        # Download Excel spreadsheet from S3
        xlsx_bytes = doc_manager.load_from_s3(spreadsheet_filename)

        with tempfile.TemporaryDirectory() as temp_dir:
            # Save Excel spreadsheet to temp file
            xlsx_path = os.path.join(temp_dir, spreadsheet_filename)
            with open(xlsx_path, 'wb') as f:
                f.write(xlsx_bytes)

            # Get sheet names from the workbook
            wb = load_workbook(xlsx_path, read_only=True)
            all_sheet_names = wb.sheetnames
            wb.close()

            # Determine which sheets to preview
            if not sheet_names:
                # Empty list means all sheets
                target_sheets = all_sheet_names
            else:
                # Validate requested sheet names
                invalid_sheets = [s for s in sheet_names if s not in all_sheet_names]
                if invalid_sheets:
                    return {
                        "content": [{
                            "text": f"**Sheet(s) not found**: {', '.join(invalid_sheets)}\n\n"
                                   f"Available sheets: {', '.join(all_sheet_names)}"
                        }],
                        "status": "error"
                    }
                target_sheets = sheet_names

            # Convert Excel to PDF using LibreOffice
            logger.info(f"Converting {spreadsheet_filename} to PDF...")
            result = subprocess.run(
                ['soffice', '--headless', '--convert-to', 'pdf', '--outdir', temp_dir, xlsx_path],
                capture_output=True,
                text=True,
                timeout=60  # 60 second timeout
            )

            if result.returncode != 0:
                logger.error(f"LibreOffice conversion failed: {result.stderr}")
                return {
                    "content": [{
                        "text": f"**PDF conversion failed**\n\n{result.stderr}"
                    }],
                    "status": "error"
                }

            pdf_path = os.path.join(temp_dir, spreadsheet_filename.replace('.xlsx', '.pdf'))

            if not os.path.exists(pdf_path):
                return {
                    "content": [{
                        "text": "**PDF file not created**\n\nLibreOffice conversion may have failed silently."
                    }],
                    "status": "error"
                }

            # Get total pages in PDF (each sheet becomes a page)
            from pdf2image import pdfinfo_from_path
            pdf_info = pdfinfo_from_path(pdf_path)
            total_pages = pdf_info.get('Pages', len(all_sheet_names))

            # Build content with images
            # Note: LibreOffice converts sheets in order, so page N = sheet N
            content = [{
                "text": f"**{spreadsheet_filename}** - {len(target_sheets)} sheet(s) of {len(all_sheet_names)} total"
            }]

            for sheet_name in target_sheets:
                # Get page number (1-indexed, sheets are in order)
                try:
                    page_num = all_sheet_names.index(sheet_name) + 1
                except ValueError:
                    continue

                if page_num > total_pages:
                    content.append({"text": f"**{sheet_name}:** (page {page_num} exceeds PDF pages)"})
                    continue

                logger.info(f"Converting sheet '{sheet_name}' (page {page_num}) to image...")
                images = convert_from_path(
                    pdf_path,
                    first_page=page_num,
                    last_page=page_num,
                    dpi=150
                )

                if images:
                    img_buffer = io.BytesIO()
                    images[0].save(img_buffer, format='PNG')
                    img_bytes = img_buffer.getvalue()

                    content.append({"text": f"**Sheet: {sheet_name}**"})
                    content.append({
                        "image": {
                            "format": "png",
                            "source": {"bytes": img_bytes}
                        }
                    })

            logger.info(f"Successfully generated {len(target_sheets)} preview(s)")

            text_blocks = [b for b in content if "text" in b]
            image_blocks = [b for b in content if "image" in b]
            return build_image_response(text_blocks, image_blocks, {
                "filename": spreadsheet_filename,
                "sheet_names": target_sheets,
                "all_sheets": all_sheet_names,
                "tool_type": "excel_spreadsheet",
                "user_id": user_id,
                "session_id": session_id,
                "hideImageInChat": True
            })

    except subprocess.TimeoutExpired:
        logger.error("LibreOffice conversion timed out")
        return {
            "content": [{
                "text": "**Conversion timed out**\n\nThe spreadsheet may be too large or complex."
            }],
            "status": "error"
        }
    except Exception as e:
        logger.error(f"preview_excel_sheets failed: {e}")
        return {
            "content": [{
                "text": f"**Failed to generate preview**\n\n{str(e)}"
            }],
            "status": "error"
        }


# --- Skill registration ---
register_skill("excel-spreadsheets", tools=[create_excel_spreadsheet, modify_excel_spreadsheet, list_my_excel_spreadsheets, read_excel_spreadsheet, preview_excel_sheets])
